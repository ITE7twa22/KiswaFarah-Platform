from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
import mysql.connector
from datetime import datetime, date, timedelta
from flask import send_file
from io import BytesIO
from openpyxl import Workbook
import pandas as pd
import io




app = Flask(__name__)
app.secret_key = "secretkey"

# الاتصال بقاعدة البيانات
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="attendance_database",
    charset="utf8mb4"
)
cursor = db.cursor(dictionary=True)

# صفحة تسجيل الدخول
@app.route("/", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        name = request.form.get("username", "").strip()
        code = request.form.get("code", "").strip()

        if not name or not code:
            error = "⚠️ يرجى إدخال الاسم والكود"
            return render_template("login.html", error=error)

        # تحقق أولاً إذا كان قائد ركن عادي
        cursor.execute("SELECT * FROM Leader WHERE name = %s AND id = %s", (name, code))
        leader = cursor.fetchone()

        if leader:
            session["username"] = leader["name"]
            session["user_type"] = "leader"
            session["section_id"] = leader["section_id"]
            session["leader_id"] = leader["id"]
            return redirect(url_for("dashboard"))

        # تحقق من قائدة قسم
        cursor.execute("SELECT * FROM DepartmentLeader WHERE name = %s AND id = %s", (name, code))
        dept_leader = cursor.fetchone()

        if dept_leader:
            session["username"] = dept_leader["name"]
            session["user_type"] = "department"
            session["leader_id"] = dept_leader["id"]
            return redirect(url_for("department_dashboard"))

        # تحقق من قائدة لجنة
        cursor.execute("SELECT * FROM CommitteeLeader WHERE name = %s AND id = %s", (name, code))
        committee_leader = cursor.fetchone()

        if committee_leader:
            session["username"] = committee_leader["name"]
            session["user_type"] = "committee"
            session["leader_id"] = committee_leader["id"]
            return redirect(url_for("committee_dashboard"))

        # إذا لم يتم العثور على أي نوع
        error = "❌ لم يتم العثور على القائدة. تأكدي من الاسم والرقم."

    return render_template("login.html", error=error)




from datetime import datetime, date
from flask import render_template, redirect, session, url_for, flash





@app.route("/update_note", methods=["POST"])
def update_note():
    if "username" not in session:
        return redirect(url_for("login"))

    note_text = request.form.get("note", "").strip()
    volunteer_index = request.form.get("row_index")

    if not note_text or volunteer_index is None:
        flash("⚠️ تأكد من إدخال الملاحظة.", "danger")
        return redirect(url_for("dashboard"))

    try:
        volunteer_index = int(volunteer_index)
    except ValueError:
        flash("⚠️ رقم المتطوعة غير صالح.", "danger")
        return redirect(url_for("dashboard"))

    # الحصول على معلومات القائدة
    leader_name = session["username"]
    cursor.execute("SELECT id, section_id FROM Leader WHERE name = %s", (leader_name,))
    leader = cursor.fetchone()

    if not leader:
        flash("⚠️ لم يتم العثور على القائدة.", "danger")
        return redirect(url_for("dashboard"))

    section_id = leader["section_id"]

    # جلب المتطوعات الخاصة بهذا الركن بنفس ترتيب الواجهة
    cursor.execute("SELECT * FROM Volunteer WHERE section_id = %s ORDER BY id", (section_id,))
    volunteers = cursor.fetchall()

    if volunteer_index < 0 or volunteer_index >= len(volunteers):
        flash("⚠️ المتطوعة غير موجودة.", "danger")
        return redirect(url_for("dashboard"))

    selected_volunteer = volunteers[volunteer_index]
    volunteer_id = selected_volunteer["id"]

    # تحديث عمود notes في جدول Volunteer
    cursor.execute("UPDATE Volunteer SET notes = %s WHERE id = %s", (note_text, volunteer_id))
    db.commit()

    flash("✅ تم تحديث الملاحظة بنجاح.", "success")
    return redirect(url_for("dashboard"))

@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("login"))

    leader_name = session["username"]

    # جلب بيانات القائدة
    cursor.execute("SELECT * FROM Leader WHERE name = %s", (leader_name,))
    leader = cursor.fetchone()
    section_id = leader["section_id"]

    # جلب اسم الركن
    cursor.execute("SELECT name FROM Section WHERE id = %s", (section_id,))
    section = cursor.fetchone()["name"]

    # جلب المتطوعين في الركن
    cursor.execute("SELECT * FROM Volunteer WHERE section_id = %s", (section_id,))
    volunteers = cursor.fetchall()

    today = date.today()
    today_display = today.isoformat()

    # جلب آخر سجل تواريخ الحملة والمعرض
    cursor.execute("SELECT * FROM CampaignDates ORDER BY created_at DESC LIMIT 1")
    campaign_dates = cursor.fetchone()

    campaign_start = campaign_dates["campaign_start_date"] if campaign_dates else None
    campaign_end = campaign_dates["campaign_end_date"] if campaign_dates else None
    exhibition_start = campaign_dates["exhibition_start_date"] if campaign_dates else None
    exhibition_end = campaign_dates["exhibition_end_date"] if campaign_dates else None

    # حساب الأيام المتبقية
    today = date.today()
    remaining_days_campaign = (campaign_end - today).days if campaign_end else None
    remaining_days_exhibition = (exhibition_end - today).days if exhibition_end else None



    # جلب كل التواريخ المميزة من جدول الحضور
    cursor.execute("""
        SELECT DISTINCT attendance_date 
        FROM volunteerattendance 
        ORDER BY attendance_date
    """)
    date_rows = cursor.fetchall()
    date_columns = [row["attendance_date"].strftime("%Y-%m-%d") for row in date_rows]

    # عدد الحاضرين اليوم
    cursor.execute("""
        SELECT COUNT(DISTINCT volunteer_id) as count
        FROM volunteerattendance
        WHERE attendance_date = %s
        AND status LIKE '✔%%'
        AND volunteer_id IN (SELECT id FROM Volunteer WHERE section_id = %s)
    """, (today, section_id))
    attended = cursor.fetchone()["count"]

    # الأيام المتبقية
    if date_columns:
        last_date = datetime.strptime(date_columns[-1], "%Y-%m-%d").date()
        remaining_days = max((last_date - today).days, 0)
        is_penultimate_day = remaining_days in [0, 1, 2]
    else:
        remaining_days = None
        is_penultimate_day = False

    # تحليل الحضور لكل متطوعة
    for v in volunteers:
        v_id = v["id"]

        cursor.execute("""
            SELECT attendance_date, status 
            FROM volunteerattendance 
            WHERE volunteer_id = %s
        """, (v_id,))
        attendance_records = cursor.fetchall()

        attendance_map = {
            r["attendance_date"].strftime("%Y-%m-%d"): r["status"]
            for r in attendance_records
        }
        # بعد إنشاء attendance_map
        v["marked_today"] = today.strftime("%Y-%m-%d") in attendance_map and '✔' in attendance_map[today.strftime("%Y-%m-%d")]


        attendance_marks = [attendance_map.get(d, '') for d in date_columns]
        present_days = sum(1 for mark in attendance_marks if '✔' in mark)
        total_days = len(date_columns)
        attendance_ratio = present_days / total_days if total_days > 0 else 0

        # حساب الانقطاع المتتالي
        absent_streak = 0
        max_streak = 0
        for mark in attendance_marks:
            if not mark or '✔' not in mark:
                absent_streak += 1
                max_streak = max(max_streak, absent_streak)
            else:
                absent_streak = 0

        eligible = "✔" if attendance_ratio >= 0.7 and max_streak < 3 else "✖"
        v["attendance_count"] = present_days
        v["eligible"] = eligible
        v["attendance_marks"] = attendance_marks

    # الملاحظات المرتبطة بالقائدة
    cursor.execute("""
        SELECT 
            Notes.note,
            Notes.created_at,
            Volunteer.name AS volunteer_name,
            Volunteer.phone AS volunteer_phone
        FROM Notes
        JOIN Volunteer ON Volunteer.id = Notes.volunteer_id
        WHERE Notes.leader_id = %s
        ORDER BY Notes.created_at DESC
    """, (leader["id"],))
    leader_notes = cursor.fetchall()

    return render_template("dashboard.html",
                           username=leader_name,
                           section=section,
                           total=len(volunteers),
                           attended=attended,
                           data=volunteers,
                           headers=["الاسم", "الجوال"] + date_columns,
                           today=today_display,
                           date_col=today.strftime("%Y-%m-%d"),
                           eligible_col="eligible",
                           notes_col="notes",
                           is_penultimate_day=is_penultimate_day,
                           leader_notes=leader_notes,
                           campaign_start=campaign_start,
                           exhibition_start=exhibition_start,
                           remaining_days=remaining_days_campaign,
                           remaining_days_exhibition=remaining_days_exhibition)

@app.route("/mark", methods=["POST"])
def mark():
    if "username" not in session:
        return redirect(url_for("login"))

    name = request.form.get("name")
    leader_name = session["username"]

    # نحاول نجيب القائدة من جدول Leader (قائدة ركن)
    cursor.execute("SELECT * FROM Leader WHERE name = %s", (leader_name,))
    leader = cursor.fetchone()

    if leader:
        section_id = leader["section_id"]
        redirect_page = "dashboard"
        leader_id = leader["id"]
        dept_leader_id = None  # ما نحتاج نحفظ department_leader_id
    else:
        # نحاول نجيبها من جدول departmentleader
        cursor.execute("SELECT * FROM departmentleader WHERE name = %s", (leader_name,))
        dept_leader = cursor.fetchone()

        if not dept_leader:
            flash("⚠️ لم يتم العثور على بيانات القائد.", "danger")
            return redirect(url_for("login"))

        # نجيب القسم من المتطوعة نفسها
        cursor.execute("SELECT section_id FROM Volunteer WHERE name = %s", (name,))
        result = cursor.fetchone()
        if not result:
            flash("⚠️ لم يتم العثور على المتطوعة.", "danger")
            return redirect(url_for("department_dashboard"))

        section_id = result["section_id"]
        redirect_page = "department_dashboard"
        dept_leader_id = dept_leader["id"]
        leader_id = None  # ما نحتاج نحفظ leader_id

    # جلب معرف المتطوعة
    cursor.execute("SELECT id FROM Volunteer WHERE name = %s AND section_id = %s", (name, section_id))
    volunteer = cursor.fetchone()
    if not volunteer:
        flash("⚠️ لم يتم العثور على المتطوعة.", "danger")
        return redirect(url_for(redirect_page))

    volunteer_id = volunteer["id"]
    today = date.today()

    # تحقق إذا فيه تحضير سابق
    cursor.execute("""
        SELECT * FROM volunteerattendance 
        WHERE volunteer_id = %s AND attendance_date = %s
    """, (volunteer_id, today))
    existing = cursor.fetchone()

    if existing:
        flash("⚠️ تم تسجيل الحضور مسبقًا لهذه المتطوعة اليوم.", "warning")
    else:
        cursor.execute("""
            INSERT INTO volunteerattendance (
                volunteer_id, attendance_date, status, leader_id, department_leader_id, created_at
            ) VALUES (%s, %s, %s, %s, %s, NOW())
        """, (volunteer_id, today, "✔", leader_id, dept_leader_id))
        db.commit()
        flash(f"✅ تم تسجيل حضور المتطوعة {name} بواسطة {leader_name}", "success")

    return redirect(url_for(redirect_page, section_id=section_id))



@app.route("/add_volunteer", methods=["POST"])
def add_volunteer():
    if "username" not in session:
        return redirect(url_for("login"))

    full_name = request.form.get("full_name", "").strip()
    phone = request.form.get("phone", "").strip()

    # تحقق من الاسم ورقم الجوال
    if len(full_name.split()) < 3:
        flash("يرجى إدخال الاسم الثلاثي على الأقل.", "danger")
        return redirect(url_for("dashboard"))
    if not (phone.startswith("05") and phone.isdigit() and len(phone) == 10):
        flash("رقم الجوال غير صحيح.", "danger")
        return redirect(url_for("dashboard"))

    # جلب بيانات القائدة
    leader_name = session["username"]
    cursor.execute("SELECT * FROM Leader WHERE name = %s", (leader_name,))
    leader = cursor.fetchone()
    section_id = leader["section_id"]

    # تحقق هل المتطوعة موجودة بأي ركن آخر
    cursor.execute("SELECT section_id FROM Volunteer WHERE name = %s", (full_name,))
    existing = cursor.fetchone()
    if existing:
        existing_section_id = existing["section_id"]
        if existing_section_id != section_id:
            # جلب اسم الركن
            cursor.execute("SELECT name FROM Section WHERE id = %s", (existing_section_id,))
            section_name = cursor.fetchone()["name"]
            flash(f"⚠️ هذه المتطوعة موجودة بالفعل في ركن: {section_name}", "warning")
            return redirect(url_for("dashboard"))

    # تحقق من عدم تكرار الاسم في نفس القسم
    cursor.execute("SELECT * FROM Volunteer WHERE name = %s AND section_id = %s", (full_name, section_id))
    if cursor.fetchone():
        flash("⚠️ المتطوعة موجودة مسبقًا في الركن.", "warning")
        return redirect(url_for("dashboard"))

    # إضافة المتطوعة
    cursor.execute("INSERT INTO Volunteer (name, phone, section_id) VALUES (%s, %s, %s)",
                   (full_name, phone, section_id))
    db.commit()

    flash("✅ تم إضافة المتطوعة بنجاح", "success")
    return redirect(url_for("dashboard"))


@app.route("/send_note", methods=["POST"])
def send_note():
    if "username" not in session:
        return redirect(url_for("login"))

    volunteer_id = request.form.get("volunteer_id")
    note_text = request.form.get("note", "").strip()
    section_id = request.args.get("section")
    selected_leaders = request.form.getlist("leaders")

    if not (volunteer_id and note_text and section_id):
        flash("⚠️ تأكد من تعبئة جميع الحقول.", "danger")
        return redirect(url_for("hr_dashboard"))

    try:
        volunteer_id = int(volunteer_id)
        section_id = int(section_id)
    except ValueError:
        flash("⚠️ بيانات غير صالحة.", "danger")
        return redirect(url_for("hr_dashboard"))

    if not selected_leaders:
        cursor.execute("SELECT id FROM Leader WHERE section_id = %s", (section_id,))
        leaders = cursor.fetchall()
        selected_leaders = [leader["id"] for leader in leaders]
    else:
        selected_leaders = [int(lid) for lid in selected_leaders]

    for leader_id in selected_leaders:
        cursor.execute("""
            INSERT INTO Notes (volunteer_id, leader_id, note, created_at)
            VALUES (%s, %s, %s, NOW())
        """, (volunteer_id, leader_id, note_text))

    db.commit()

    flash(f"📌 تم إرسال الملاحظة إلى {len(selected_leaders)} قائد(ة).", "success")
    return redirect(url_for("hr_dashboard", section=section_id))


@app.route("/transfer-request", methods=["GET"])
def transfer_request():
    name = request.args.get("volunteer_name")
    phone = request.args.get("volunteer_phone")

    # جلب بيانات المتطوعة
    cursor.execute("SELECT * FROM Volunteer WHERE name = %s AND phone = %s", (name, phone))
    volunteer = cursor.fetchone()

    if not volunteer:
        flash("⚠️ لم يتم العثور على المتطوعة.", "danger")
        return redirect(url_for("dashboard"))

    # جلب اسم الركن الحالي
    cursor.execute("SELECT name FROM Section WHERE id = %s", (volunteer["section_id"],))
    from_section = cursor.fetchone()["name"]

    # جلب كل أسماء الأركان الأخرى
    cursor.execute("SELECT name FROM Section WHERE id != %s", (volunteer["section_id"],))
    other_sections = [row["name"] for row in cursor.fetchall()]

    return render_template("transfer_form.html",
                           volunteer_name=name,
                           volunteer_phone=phone,
                           from_section=from_section,
                           all_sections=other_sections)


@app.route("/submit-transfer", methods=["POST"])
def submit_transfer():
    if "username" not in session:
        return redirect(url_for("login"))

    leader_name = session["username"]
    volunteer_name = request.form.get("volunteer_name")
    volunteer_phone = request.form.get("volunteer_phone")
    from_section_name = request.form.get("from_section")
    to_section_name = request.form.get("to_section")
    note = request.form.get("notes", "")
    today = datetime.now()


    # جلب معرف الركن الحالي والمستهدف
    cursor.execute("SELECT id FROM Section WHERE name = %s", (from_section_name,))
    from_section_id = cursor.fetchone()["id"]
    cursor.execute("SELECT id FROM Section WHERE name = %s", (to_section_name,))
    to_section_id = cursor.fetchone()["id"]

    # جلب بيانات المتطوعة
    cursor.execute("SELECT * FROM Volunteer WHERE name = %s AND phone = %s AND section_id = %s",
                   (volunteer_name, volunteer_phone, from_section_id))
    volunteer = cursor.fetchone()

    if not volunteer:
        flash("⚠️ لم يتم العثور على المتطوعة.", "danger")
        return redirect(url_for("dashboard"))

    if volunteer["number_of_requests"] >= 1:
        flash(f"⚠️ المتطوعة {volunteer_name} تجاوزت الحد المسموح به (محاولة واحدة).", "danger")
        return redirect(url_for("dashboard"))

    # جلب معرف القائد
    cursor.execute("SELECT id FROM Leader WHERE name = %s", (leader_name,))
    leader = cursor.fetchone()

    # إضافة الطلب
    cursor.execute("""
        INSERT INTO Request (leader_id, volunteer_id, request_date, from_section_id, to_section_id, note)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (leader["id"], volunteer["id"], today, from_section_id, to_section_id, note))
    db.commit()

    # تحديث عدد محاولات المتطوعة
    cursor.execute("UPDATE Volunteer SET number_of_requests = number_of_requests + 1 WHERE id = %s", (volunteer["id"],))
    db.commit()

    flash("✅ تم إرسال طلب النقل بنجاح.", "success")
    return redirect(url_for("dashboard"))


@app.route("/transfer_requests")
def transfer_requests():
    cursor.execute("""
        SELECT r.*, v.name AS volunteer, v.phone, v.number_of_requests AS attempts,
               l.name AS leader, s1.name AS from_section, s2.name AS to_section
        FROM Request r
        JOIN Volunteer v ON r.volunteer_id = v.id
        JOIN Leader l ON r.leader_id = l.id
        JOIN Section s1 ON r.from_section_id = s1.id
        JOIN Section s2 ON r.to_section_id = s2.id
        ORDER BY r.request_date DESC
    """)
    requests = cursor.fetchall()

    pending = [r for r in requests if r["status"] == "جاري المعالجة"]
    completed = [r for r in requests if r["status"] in ["تم النقل", "مرفوض"]]

    # جلب الحد الأدنى والأعلى والفِعلي لكل ركن
    cursor.execute("SELECT id, name, min, max FROM Section")
    sections = cursor.fetchall()

    section_limits = {}
    for sec in sections:
        cursor.execute("SELECT COUNT(*) as count FROM Volunteer WHERE section_id = %s", (sec["id"],))
        current_count = cursor.fetchone()["count"]
        section_limits[sec["name"]] = {
        "min": sec["min"],
        "max": sec["max"],
        "current": current_count
    }


    return render_template("transfer_requests.html",
                           pending_requests=pending,
                           completed_requests=completed,
                           status_counts={
                               "قيد المراجعة": len(pending),
                               "تم النقل": len([r for r in requests if r["status"] == "تم النقل"]),
                               "مرفوض": len([r for r in requests if r["status"] == "مرفوض"]),
                           },
                           section_limits=section_limits)


@app.route("/process_transfer", methods=["POST"])
def process_transfer():
    if "username" not in session:
        return redirect(url_for("login"))

    request_id = request.form.get("request_id")
    action = request.form.get("action")  # 'accept' or 'reject'
    reason = request.form.get("reason", "").strip()

    # جلب الطلب
    cursor.execute("SELECT * FROM Request WHERE id = %s", (request_id,))
    transfer = cursor.fetchone()

    if not transfer:
        flash("⚠️ لم يتم العثور على الطلب.", "danger")
        return redirect(url_for("transfer_requests"))

    if action == "accept":
        # 1. نقل المتطوعة للركن الجديد
        cursor.execute("UPDATE Volunteer SET section_id = %s WHERE id = %s",
                       (transfer["to_section_id"], transfer["volunteer_id"]))
        db.commit()

        # 2. تحديث حالة الطلب
        cursor.execute("UPDATE Request SET status = 'تم النقل', note = '—' WHERE id = %s", (request_id,))
        db.commit()

        flash("✅ تم نقل المتطوعة بنجاح.", "success")

    elif action == "reject":
        # 1. تحديث حالة الطلب مع السبب
        cursor.execute("UPDATE Request SET status = 'مرفوض', note = %s WHERE id = %s", (reason, request_id))
        db.commit()

        flash("❌ تم رفض الطلب.", "warning")

    return redirect(url_for("transfer_requests"))


from datetime import datetime
from flask import request, redirect, url_for, flash, session, render_template

@app.route("/hr")
def hr_dashboard():
    if "username" not in session:
        return redirect(url_for("login"))
    
    cursor.execute("SELECT * FROM CampaignDates ORDER BY created_at DESC LIMIT 1")
    latest_campaign = cursor.fetchone()


    section_id = request.args.get("section")

    # جلب كل الأركان
    cursor.execute("SELECT * FROM Section")
    sections = cursor.fetchall()
    

    selected_section = None
    section_data = []
    section_leaders = []
    today = datetime.today().date()
    today_col = today.strftime("%Y_%m_%d")
    today_display = today.isoformat()
    attended = total = 0
    remaining_days = None
    top_sections = []
    hr_chart_labels = []
    hr_chart_data = []

    # جلب أعمدة الحضور من الجدول
    cursor.execute("SHOW COLUMNS FROM Volunteer")
    all_columns_data = cursor.fetchall()
    all_columns = [col["Field"] for col in all_columns_data]
    date_columns = [col for col in all_columns if col.startswith("2025_")]
    cursor.execute("SELECT COUNT(DISTINCT committee_section) AS total_committees FROM committeeleader")
    committee_leader_count = cursor.fetchone()["total_committees"]


    # ✅ تأكد أن today_col موجود بين الأعمدة
    if today_col not in all_columns:
        today_col = None

    if section_id:
        cursor.execute("SELECT * FROM Section WHERE id = %s", (section_id,))
        selected_section = cursor.fetchone()

        if selected_section:
            cursor.execute("SELECT id, name FROM Leader WHERE section_id = %s", (section_id,))
            section_leaders = cursor.fetchall()

            cursor.execute("SELECT * FROM Volunteer WHERE section_id = %s", (section_id,))
            raw_volunteers = cursor.fetchall()

            total = len(raw_volunteers)
            section_data = []

            for volunteer in raw_volunteers:
                attendance_days = 0
                consecutive_absent = 0
                max_consecutive_absent = 0

                for date_col in date_columns:
                    status = volunteer.get(date_col) or ""
                    if status.startswith("✔"):
                        attendance_days += 1
                        consecutive_absent = 0
                    else:
                        consecutive_absent += 1
                        max_consecutive_absent = max(max_consecutive_absent, consecutive_absent)

                total_days = len(date_columns)
                percentage = (attendance_days / total_days) * 100 if total_days > 0 else 0

                if today_col:
                    value = volunteer.get(today_col)
                    present_today = "✔" if value and value.startswith("✔") else "❌"
                else:
                    present_today = "❌"

                is_candidate = "✔" if percentage >= 70 and max_consecutive_absent < 4 else "❌"

                section_data.append({
                    "id": volunteer["id"],
                    "name": volunteer["name"],
                    "phone": volunteer["phone"],
                    "attendance_days": attendance_days,
                    "attendance_percentage": f"{percentage:.0f}%",
                    "present_today": present_today,
                    "candidate": is_candidate,
                    "note": volunteer.get("note", "")
                })

            if date_columns:
                last_date = max(datetime.strptime(d, "%Y_%m_%d") for d in date_columns)
                remaining_days = max((last_date.date() - today).days, 0)

    else:
        top_ratio = 0

        for section in sections:
            # ✅ نفذ الاستعلام فقط إذا اليوم موجود ضمن الأعمدة
            if today_col:
                cursor.execute(
                    f"SELECT COUNT(*) AS count FROM Volunteer WHERE section_id = %s AND `{today_col}` LIKE '✔%%'",
                    (section["id"],)
                )
                attended_today = cursor.fetchone()["count"]
            else:
                attended_today = 0

            cursor.execute("SELECT COUNT(*) AS total FROM Volunteer WHERE section_id = %s", (section["id"],))
            total_count = cursor.fetchone()["total"]

            if total_count > 0:
                ratio = attended_today / total_count
                if ratio > top_ratio:
                    top_ratio = ratio
                    top_sections = [section["name"]]
                elif ratio == top_ratio:
                    top_sections.append(section["name"])

            hr_chart_labels.append(section["name"])
            hr_chart_data.append(attended_today)

    return render_template("hr.html",
                           all_sections=sections,
                           selected_section=selected_section,
                           section_data=section_data,
                           section_leaders=section_leaders,
                           today=today_display,
                           total=total,
                           attended=attended,
                           remaining_days=remaining_days,
                           latest_campaign=latest_campaign,
                           committee_leader_count=committee_leader_count,
                           hr_chart_labels=hr_chart_labels if not section_id else None,
                           hr_chart_data=hr_chart_data if not section_id else None,
                           top_section_name=", ".join(top_sections) if not section_id else None)



@app.route("/mark_attendance", methods=["POST"])
def mark_attendance():
    volunteer_id = request.form.get("volunteer_id")
    today_col = date.today().strftime("%Y_%m_%d")

    # تأكد من أن العمود موجود
    cursor.execute("SHOW COLUMNS FROM Volunteer")
    columns = [col["Field"] for col in cursor.fetchall()]
    if today_col not in columns:
        cursor.execute(f"ALTER TABLE Volunteer ADD `{today_col}` VARCHAR(10)")
        db.commit()

    cursor.execute(f"UPDATE Volunteer SET `{today_col}` = %s WHERE id = %s", ("✔", volunteer_id))
    db.commit()

    flash("✅ تم تسجيل الحضور", "success")
    return redirect(url_for("dashboard"))



@app.route("/department_dashboard", methods=["GET", "POST"])
def department_dashboard():
    if "username" not in session:
        return redirect(url_for("login"))

    leader_name = session["username"]

    cursor.execute("SELECT * FROM departmentleader WHERE name = %s", (leader_name,))
    dept_leader = cursor.fetchone()

    if not dept_leader:
        flash("⚠️ لا يوجد قائد قسم بهذا الاسم", "danger")
        return redirect(url_for("login"))

    dept_leader_id = dept_leader["id"]

    cursor.execute("SELECT * FROM Section WHERE department_leader_id = %s", (dept_leader_id,))
    sections = cursor.fetchall()

    section_data = []
    for sec in sections:
        cursor.execute("SELECT COUNT(*) AS count FROM Volunteer WHERE section_id = %s", (sec["id"],))
        count = cursor.fetchone()["count"]
        section_data.append({
            "id": sec["id"],
            "name": sec["name"],
            "min": sec["min"],
            "max": sec["max"],
            "count": count
        })

    cursor.execute("""
        SELECT L.*, S.name AS section_name 
        FROM Leader L
        JOIN Section S ON L.section_id = S.id
        WHERE S.department_leader_id = %s
    """, (dept_leader_id,))
    leaders = cursor.fetchall()

    selected_section_id = request.args.get("section_id")
    selected_section = None
    section_leaders = []
    section_volunteers = []
    today = date.today()
    today_col = today.strftime("%Y-%m-%d")

    # جلب كل التواريخ الفريدة من جدول الحضور
    cursor.execute("SELECT DISTINCT attendance_date FROM volunteerattendance ORDER BY attendance_date")
    date_rows = cursor.fetchall()
    date_columns = [row["attendance_date"].strftime("%Y-%m-%d") for row in date_rows]

    if selected_section_id:
        try:
            selected_section_id = int(selected_section_id)

            cursor.execute("SELECT * FROM Section WHERE id = %s", (selected_section_id,))
            selected_section = cursor.fetchone()

            cursor.execute("SELECT * FROM Leader WHERE section_id = %s", (selected_section_id,))
            section_leaders = cursor.fetchall()

            cursor.execute("SELECT * FROM Volunteer WHERE section_id = %s", (selected_section_id,))
            section_volunteers = cursor.fetchall()

            for v in section_volunteers:
                v_id = v["id"]

                # جلب حضور المتطوعة من جدول volunteerattendance
                cursor.execute("""
                    SELECT attendance_date, status 
                    FROM volunteerattendance 
                    WHERE volunteer_id = %s
                """, (v_id,))
                attendance_records = cursor.fetchall()

                attendance_map = {
                    r["attendance_date"].strftime("%Y-%m-%d"): r["status"]
                    for r in attendance_records
                }

                attendance_marks = [attendance_map.get(d, '') for d in date_columns]
                present_days = sum(1 for mark in attendance_marks if '✔' in mark)
                total_days = len(date_columns)
                attendance_ratio = present_days / total_days if total_days else 0

                # غياب 3 أيام متتالية
                absent_streak = 0
                max_streak = 0
                for mark in attendance_marks:
                    if not mark or '✔' not in mark:
                        absent_streak += 1
                        max_streak = max(max_streak, absent_streak)
                    else:
                        absent_streak = 0

                eligible = "✔" if attendance_ratio >= 0.7 and max_streak < 3 else "✖"
                v["attendance_count"] = present_days
                v["eligible"] = eligible
                v["date_columns"] = date_columns
                v["attendance_marks"] = attendance_marks
                v["marked_today"] = today_col in attendance_map and '✔' in attendance_map[today_col]

        except ValueError:
            flash("⚠️ القسم المحدد غير صالح", "danger")

    # 🟦 بيانات توزيع المتطوعات في الأركان التابعة
    leader_sections = []
    for sec in sections:
        cursor.execute("SELECT COUNT(*) AS count FROM Volunteer WHERE section_id = %s", (sec["id"],))
        count = cursor.fetchone()["count"]
        leader_sections.append({
            "name": sec["name"],
            "actual_count": count,
            "min_required": sec["min"],
            "max_allowed": sec["max"]
        })

    return render_template("department_dashboard.html",
                           username=leader_name,
                           section_data=section_data,
                           leaders=leaders,
                           selected_section=selected_section,
                           section_leaders=section_leaders,
                           section_volunteers=section_volunteers or [],
                           today_col=today_col,
                           date_columns=date_columns,
                           leader_sections=leader_sections)


from flask import jsonify  # تأكد أنك مستوردها

@app.route("/save_note", methods=["POST"])
def save_note():
    data = request.get_json()
    volunteer_id = data.get("volunteer_id")
    note = data.get("note", "").strip()

    if not volunteer_id:
        return jsonify({"status": "error", "message": "Missing volunteer_id"}), 400

    try:
        cursor.execute("UPDATE Volunteer SET notes = %s WHERE id = %s", (note, volunteer_id))
        db.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("DB Error:", e)
        return jsonify({"status": "error", "message": "Database error"}), 500


@app.route("/evaluate_leader/<int:leader_id>", methods=["GET", "POST"])
def evaluate_leader(leader_id):
    if "username" not in session:
        return redirect(url_for("login"))

    questions = [
        "مدى التزام القائد بالحضور في الوقت المحدد",
        "تعامل القائد مع المتطوعات باحترام ولباقة",
        "قدرة القائد على حل المشكلات الميدانية",
        "وضوح القائد في شرح المهام للمتطوعات",
        "سرعة استجابة القائد للتوجيهات والتعليمات",
        "تنظيم القائد للمهام اليومية",
        "متابعة القائد لأداء المتطوعات",
        "تشجيع القائد للمتطوعات وتحفيزهن",
        "مدى مساهمة القائد في تحسين بيئة العمل"
    ]

    # جلب بيانات القائد
    cursor.execute("SELECT * FROM Leader WHERE id = %s", (leader_id,))
    leader = cursor.fetchone()
    if not leader:
        flash("⚠️ لم يتم العثور على القائد.", "danger")
        return redirect(url_for("department_dashboard"))

    if request.method == "POST":
        evaluator_name = session["username"]

        # جلب evaluator_id من جدول departmentleader
        cursor.execute("SELECT id FROM departmentleader WHERE name = %s", (evaluator_name,))
        evaluator = cursor.fetchone()
        if not evaluator:
            return redirect(url_for("department_dashboard"))

        evaluator_id = evaluator["id"]

        # استخراج إجابات الأسئلة من النموذج
        answers = []
        for i in range(1, 10):
            val = request.form.get(f"q{i}")
            if not val or not val.isdigit():
                flash(f"⚠️ السؤال رقم {i} لم يتم تعبئته بشكل صحيح.", "danger")
                return redirect(request.url)
            answers.append(int(val))

        comments = request.form.get("comments", "").strip()

        # إدخال التقييم في قاعدة البيانات
        cursor.execute("""
            INSERT INTO leader_evaluations 
            (evaluator_id, leader_id, q1, q2, q3, q4, q5, q6, q7, q8, q9, comments)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (evaluator_id, leader_id, *answers, comments))
        db.commit()

        flash("✅ تم إرسال التقييم بنجاح.", "success")
        return redirect(url_for("department_dashboard"))

    return render_template("evaluate_leader.html", leader=leader, questions=questions)


@app.route("/add_volunteer_by_dept_leader", methods=["POST"])
def add_volunteer_by_dept_leader():
    if "username" not in session:
        return redirect(url_for("login"))

    full_name = request.form.get("full_name", "").strip()
    phone = request.form.get("phone", "").strip()
    section_id = request.form.get("section_id", "").strip()

    # تحقق من صحة البيانات
    if len(full_name.split()) < 3:
        flash("يرجى إدخال الاسم الثلاثي على الأقل.", "danger")
        return redirect(request.referrer or url_for("department_dashboard"))

    if not (phone.startswith("05") and phone.isdigit() and len(phone) == 10):
        flash("رقم الجوال غير صحيح.", "danger")
        return redirect(request.referrer or url_for("department_dashboard"))

    # تحقق من أن المستخدم هو قائد قسم
    leader_name = session["username"]
    cursor.execute("SELECT * FROM DepartmentLeader WHERE name = %s", (leader_name,))
    dept_leader = cursor.fetchone()

    if not dept_leader:
        flash("⚠️ لا تملك صلاحية إضافة متطوعات.", "danger")
        return redirect(url_for("login"))

    # تحقق من صحة section_id
    try:
        section_id = int(section_id)
    except ValueError:
        flash("⚠️ معرف الركن غير صالح.", "danger")
        return redirect(request.referrer or url_for("department_dashboard"))

    # تحقق هل المتطوعة موجودة في أي ركن آخر
    cursor.execute("SELECT section_id FROM Volunteer WHERE name = %s", (full_name,))
    existing = cursor.fetchone()
    if existing:
        existing_section_id = existing["section_id"]
        if existing_section_id != section_id:
            cursor.execute("SELECT name FROM Section WHERE id = %s", (existing_section_id,))
            section_name = cursor.fetchone()["name"]
            flash(f"⚠️ هذه المتطوعة موجودة بالفعل في ركن: {section_name}", "warning")
            return redirect(request.referrer or url_for("department_dashboard"))

    # تحقق من عدم تكرار الاسم في نفس الركن
    cursor.execute("SELECT * FROM Volunteer WHERE name = %s AND section_id = %s", (full_name, section_id))
    if cursor.fetchone():
        flash("⚠️ المتطوعة موجودة مسبقًا في هذا الركن.", "warning")
        return redirect(request.referrer or url_for("department_dashboard"))

    # إضافة المتطوعة
    cursor.execute("INSERT INTO Volunteer (name, phone, section_id) VALUES (%s, %s, %s)",
                   (full_name, phone, section_id))
    db.commit()

    flash("✅ تم إضافة المتطوعة بنجاح.", "success")
    return redirect(request.referrer or url_for("department_dashboard"))


from datetime import datetime
@app.route("/set_campaign_dates", methods=["POST"])
def set_campaign_dates():
    if "username" not in session:
        return redirect(url_for("login"))

    campaign_start = request.form.get("campaign_start_date")
    campaign_end = request.form.get("campaign_end_date")
    exhibition_start = request.form.get("exhibition_start_date")
    exhibition_end = request.form.get("exhibition_end_date")

    try:
        cursor.execute("""
            INSERT INTO CampaignDates (
                campaign_start_date, campaign_end_date,
                exhibition_start_date, exhibition_end_date
            )
            VALUES (%s, %s, %s, %s)
        """, (campaign_start, campaign_end, exhibition_start, exhibition_end))
        db.commit()
        flash("✅ تم حفظ جميع تواريخ الحملة والمعرض بنجاح.", "success")
    except Exception as e:
        db.rollback()
        flash(f"⚠️ حدث خطأ أثناء الحفظ: {str(e)}", "danger")

    return redirect(url_for("hr_dashboard"))


@app.route("/evaluate_volunteer", methods=["GET", "POST"])
def evaluate_volunteer():
    if "username" not in session:
        return redirect(url_for("login"))

    volunteer_id = request.args.get("volunteer_id")

    # التأكد من وجود المتطوعة
    cursor.execute("SELECT * FROM Volunteer WHERE id = %s", (volunteer_id,))
    volunteer = cursor.fetchone()
    if not volunteer:
        flash("⚠️ المتطوعة غير موجودة.", "danger")
        return redirect(url_for("dashboard"))

    # التأكد من عدم تقييمها مسبقًا
    cursor.execute("SELECT COUNT(*) AS count FROM quality WHERE volunteer_id = %s", (volunteer_id,))
    existing = cursor.fetchone()["count"]
    if existing > 0:
        flash("⚠️ تم تقييم هذه المتطوعة مسبقًا.", "warning")
        return redirect(url_for("dashboard"))

    # الأسئلة
    questions = [
        "حسن التعامل، وتقديم المساعدة عند الحاجة.",
        "الالتزام بالحضور والتواجد في الأوقات المحددة بالركن.",
        "القدرة على العمل الجماعي وحل المشاكل بشكل فعال.",
        "مستوى المعرفة بالمهام المسندة لهم واستجابة المتطوعين للتوجيهات.",
        "الالتزام بميثاق الأخلاقي.",
        "لدى المتطوع صفات قيادية؟"
    ]

    if request.method == "POST":
        leader_name = session["username"]
        cursor.execute("SELECT id FROM Leader WHERE name = %s", (leader_name,))
        leader = cursor.fetchone()
        if not leader:
            flash("⚠️ القائد غير موجود.", "danger")
            return redirect(url_for("dashboard"))

        leader_id = leader["id"]

        # جمع التقييمات
        ratings = []
        for i in range(1, 7):  # q1 إلى q6
            rating = request.form.get(f"q{i}")
            if not rating:
                flash("⚠️ تأكد من تعبئة جميع التقييمات.", "danger")
                return redirect(request.url)
            ratings.append(int(rating))

        # الملاحظات
        comments = request.form.get("comments", "").strip()

        # الإدخال في جدول الجودة
        cursor.execute("""
            INSERT INTO quality (leader_id, volunteer_id, q1, q2, q3, q4, q5, q6, comments)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (leader_id, volunteer_id, *ratings, comments))

        db.commit()
        flash("✅ تم حفظ التقييم بنجاح.", "success")
        return redirect(url_for("dashboard"))

    return render_template("evaluate_volunteer.html", volunteer=volunteer, questions=questions)


@app.route("/committee_dashboard")
def committee_dashboard():
    if "username" not in session or session.get("user_type") != "committee":
        return redirect(url_for("login"))

    leader_id = session["leader_id"]
    today = date.today()


    # اجلب بيانات القائدة
    cursor.execute("SELECT committee_section FROM CommitteeLeader WHERE id = %s", (leader_id,))
    leader = cursor.fetchone()
    if not leader:
        flash("⚠️ لم يتم العثور على بيانات القائدة.", "danger")
        return redirect(url_for("login"))

    section = leader["committee_section"]

    # اجلب تواريخ الحملة والمعرض
    cursor.execute("SELECT * FROM campaigndates ORDER BY created_at DESC LIMIT 1")
    dates = cursor.fetchone()
    if not dates:
        flash("⚠️ لم يتم العثور على تواريخ الحملة.", "danger")
        return redirect(url_for("login"))

    campaign_start = dates["campaign_start_date"]
    campaign_end = dates["campaign_end_date"]
    exhibition_start = dates["exhibition_start_date"]
    exhibition_end = dates["exhibition_end_date"]

    remaining_days = (campaign_end - today).days
    remaining_days_exhibition = (exhibition_end - today).days

    # عدد المتطوعات باللجنة
    cursor.execute("SELECT COUNT(*) AS total FROM CommitteeVolunteer WHERE section = %s", (section,))
    total = cursor.fetchone()["total"]

    # عدد الحضور اليوم
    cursor.execute("""
        SELECT COUNT(*) AS attended
        FROM CommitteeAttendance ca
        JOIN CommitteeVolunteer cv ON ca.volunteer_id = cv.id
        WHERE ca.attendance_date = %s AND cv.section = %s
    """, (today, section))
    attended = cursor.fetchone()["attended"]

    # جلب بيانات المتطوعات مع معلوماتهم من جدول Volunteer وحالة التحضير
    cursor.execute("""
        SELECT cv.id, cv.name, cv.phone,
            EXISTS (
                SELECT 1 FROM CommitteeAttendance ca
                WHERE ca.volunteer_id = cv.id AND ca.attendance_date = %s
            ) AS marked_today
        FROM CommitteeVolunteer cv
        WHERE cv.leader_id = %s
    """, (today, leader_id))

    data = cursor.fetchall()

    return render_template("committee_dashboard.html",
        username=session["username"],
        today=today,
        total=total,
        attended=attended,
        campaign_start=campaign_start,
        remaining_days=remaining_days,
        exhibition_start=exhibition_start,
        remaining_days_exhibition=remaining_days_exhibition,
        data=data
    )


@app.route("/evaluate_volunteer/<int:volunteer_id>")
def evaluate_volunteer_form(volunteer_id):
    if "username" not in session or session.get("user_type") != "committee":
        return redirect(url_for("login"))

    cursor.execute("SELECT name FROM CommitteeVolunteer WHERE id = %s", (volunteer_id,))
    volunteer = cursor.fetchone()
    if not volunteer:
        flash("⚠️ لم يتم العثور على المتطوعة.", "danger")
        return redirect(url_for("committee_dashboard"))

    questions = [
        "ما مدى التزام المتطوعة بالحضور والانصراف في الوقت المحدد؟",
        "ما مدى التزام المتطوعة بالمهام الموكلة إليها؟",
        "ما مدى تعاون المتطوعة مع بقية الفريق داخل اللجنة؟",
        "ما مدى احترام المتطوعة للتعليمات والتوجيهات الصادرة من القائدة؟",
        "ما مدى مرونة المتطوعة في التعامل مع التغييرات أو الظروف الطارئة؟",
        "ما مدى حُسن تواصل المتطوعة مع القائدة وبقية الفريق؟",
        "ما مدى حماس ومبادرة المتطوعة في أداء مهامها؟",
        "ما مدى التزام المتطوعة بالزي المناسب والسلوكيات العامة؟",
        "ما مدى قدرة المتطوعة على أداء المهام بجودة واهتمام؟",
        "ما مدى التزام المتطوعة بالأخلاقيات واحترام الآخرين أثناء العمل؟"
    ]


    return render_template("volunteer_rating.html", 
        volunteer_id=volunteer_id,
        volunteer_name=volunteer["name"],
        questions=questions
    )




@app.route("/submit_volunteer_evaluation", methods=["POST"])
def submit_volunteer_evaluation():
    if "username" not in session or session.get("user_type") != "committee":
        return redirect(url_for("login"))

    evaluator_id = session["leader_id"]
    volunteer_id = request.form.get("volunteer_id")
    comments = request.form.get("comments")
    created_at = date.today()

    answers = []
    for i in range(1, 11):
        answers.append(int(request.form.get(f"q{i}", 0)))

    cursor.execute("""
        INSERT INTO committee_volunteer_evaluations (
            evaluator_id, volunteer_id, q1, q2, q3, q4, q5, q6, q7, q8, q9, q10, comments, created_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (evaluator_id, volunteer_id, *answers, comments, created_at))

    db.commit()
    flash("✅ تم حفظ التقييم بنجاح.", "success")
    return redirect(url_for("committee_dashboard"))



@app.route("/committee_mark", methods=["POST"])
def committee_mark():
    if "username" not in session or session.get("user_type") != "committee":
        return redirect(url_for("login"))

    name = request.form.get("name")
    leader_id = session["leader_id"]

    # نحصل على volunteer_id
    cursor.execute("SELECT id FROM committeevolunteer WHERE name = %s", (name,))
    volunteer = cursor.fetchone()
    if not volunteer:
        flash("⚠️ المتطوعة غير موجودة.", "danger")
        return redirect(url_for("committee_dashboard"))

    volunteer_id = volunteer["id"]

    # التأكد أنه ما تم تحضيرها اليوم
    cursor.execute("""
        SELECT COUNT(*) AS count FROM CommitteeAttendance
        WHERE volunteer_id = %s AND attendance_date = CURDATE()
    """, (volunteer_id,))
    if cursor.fetchone()["count"] > 0:
        flash("⚠️ تم تحضير المتطوعة مسبقًا.", "warning")
        return redirect(url_for("committee_dashboard"))

    # إضافة الحضور
    cursor.execute("""
        INSERT INTO CommitteeAttendance (volunteer_id, attendance_date, status, leader_id, created_at)
        VALUES (%s, CURDATE(), 'present', %s, NOW())
    """, (volunteer_id, leader_id))
    db.commit()

    flash("✅ تم تحضير المتطوعة بنجاح.", "success")
    return redirect(url_for("committee_dashboard"))

@app.route("/add_committee_volunteer", methods=["POST"])
def add_committee_volunteer():
    if "username" not in session or session.get("user_type") != "committee":
        return redirect(url_for("login"))

    name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    leader_id = session.get("leader_id")

    # نجيب سكشن القائدة الحالية
    cursor.execute("SELECT committee_section FROM CommitteeLeader WHERE id = %s", (leader_id,))
    result = cursor.fetchone()

    if not result:
        flash("⚠️ لم يتم العثور على بيانات القائدة.", "danger")
        return redirect(url_for("committee_dashboard"))

    section = result["committee_section"]

    if not name or not phone:
        flash("⚠️ يرجى إدخال الاسم ورقم الجوال.", "danger")
        return redirect(url_for("committee_dashboard"))

    cursor.execute("""
        INSERT INTO CommitteeVolunteer (name, phone, section, leader_id, created_at)
        VALUES (%s, %s, %s, %s, NOW())
    """, (name, phone, section, leader_id))

    db.commit()
    flash("✅ تم إضافة المتطوعة بنجاح.", "success")
    return redirect(url_for("committee_dashboard"))

@app.route("/quality_dashboard")
def quality_dashboard():
    return render_template("quality_dashboard.html")

@app.route("/quality_field")
def quality_field():
    # جلب قائمة الأركان للفلترة
    cursor.execute("SELECT id, name FROM section")
    sections = cursor.fetchall()

    # استلام الفلاتر
    section_id = request.args.get("section_id")
    question = request.args.get("question")  # q1, q2, ...
    answer = request.args.get("answer")

    # بناء الاستعلام الأساسي
    query = """
        SELECT 
            v.name AS volunteer_name,
            v.phone AS volunteer_phone,
            l.name AS leader_name,
            s.name AS section_name,
            q.q1, q.q2, q.q3, q.q4, q.q5, q.q6,
            q.comments,
            q.created_at
        FROM quality q
        JOIN volunteer v ON q.volunteer_id = v.id
        JOIN leader l ON q.leader_id = l.id
        JOIN section s ON v.section_id = s.id
        WHERE 1=1
    """
    values = []

    if section_id:
        query += " AND s.id = %s"
        values.append(section_id)

    if question and question in ["q1", "q2", "q3", "q4", "q5", "q6"]:
        if answer and answer.isdigit():
            query += f" AND q.{question} = %s"
            values.append(int(answer))

    query += " ORDER BY q.created_at DESC"

    cursor.execute(query, values)
    evaluations = cursor.fetchall()

    return render_template("quality_field.html",
        evaluations=evaluations,
        sections=sections,
        selected_section=int(section_id) if section_id else "",
        selected_question=question,
        selected_answer=int(answer) if answer and answer.isdigit() else ""
    )

@app.route("/quality_leader")
def quality_leader():
    # جلب قائمة الأركان
    cursor.execute("SELECT id, name FROM section")
    sections = cursor.fetchall()

    # استلام فلاتر
    section_id = request.args.get("section_id")
    question = request.args.get("question")  # مثل: q4
    answer = request.args.get("answer")

    # بناء الاستعلام
    query = """
        SELECT 
            dl.name AS evaluator_name,
            l.name AS leader_name,
            l.phone AS leader_phone,
            s.name AS section_name,
            e.q1, e.q2, e.q3, e.q4, e.q5, e.q6, e.q7, e.q8, e.q9,
            e.comments,
            e.created_at
        FROM leader_evaluations e
        JOIN departmentleader dl ON e.evaluator_id = dl.id
        JOIN leader l ON e.leader_id = l.id
        JOIN section s ON l.section_id = s.id
        WHERE 1=1
    """
    values = []

    # فلتر القسم
    if section_id:
        query += " AND s.id = %s"
        values.append(section_id)

    # فلتر سؤال وإجابة
    if question and question in [f"q{i}" for i in range(1, 10)]:
        if answer and answer.isdigit():
            query += f" AND e.{question} = %s"
            values.append(int(answer))

    query += " ORDER BY e.created_at DESC"

    cursor.execute(query, values)
    evaluations = cursor.fetchall()

    return render_template("quality_leader.html",
        evaluations=evaluations,
        sections=sections,
        selected_section=int(section_id) if section_id else "",
        selected_question=question,
        selected_answer=int(answer) if answer and answer.isdigit() else ""
    )


@app.route("/quality_committee")
def quality_committee():
    # جلب جميع أسماء اللجان (committee_section) من القائدات
    cursor.execute("SELECT DISTINCT committee_section FROM committeeleader")
    sections_raw = cursor.fetchall()
    sections = [row["committee_section"] for row in sections_raw]

    # استلام الفلاتر
    section_id = request.args.get("section_id")
    question = request.args.get("question")
    answer = request.args.get("answer")

    # الاستعلام الأساسي
    query = """
        SELECT 
            cl.name AS evaluator_name,
            cv.name AS volunteer_name,
            cv.phone AS volunteer_phone,
            cl.committee_section AS section_name,
            e.q1, e.q2, e.q3, e.q4, e.q5, e.q6, e.q7, e.q8, e.q9, e.q10,
            e.comments,
            e.created_at
        FROM committee_volunteer_evaluations e
        JOIN committeeleader cl ON e.evaluator_id = cl.id
        JOIN committeevolunteer cv ON e.volunteer_id = cv.id
        WHERE 1=1
    """
    values = []

    # فلتر حسب اللجنة
    if section_id:
        query += " AND cl.committee_section = %s"
        values.append(section_id)

    # فلتر حسب سؤال وإجابة
    if question and question in [f"q{i}" for i in range(1, 11)]:
        if answer and answer.isdigit():
            query += f" AND e.{question} = %s"
            values.append(int(answer))

    query += " ORDER BY e.created_at DESC"

    cursor.execute(query, values)
    evaluations = cursor.fetchall()

    return render_template("quality_committee.html",
        evaluations=evaluations,
        sections=sections,
        selected_section=section_id,
        selected_question=question,
        selected_answer=int(answer) if answer and answer.isdigit() else ""
    )



@app.route("/download_quality_excel")
def download_quality_excel():
    wb = Workbook()
    
    # -------------------- 1. متطوعي الميدان --------------------
    ws1 = wb.active
    ws1.title = "متطوعي الميدان"

    cursor.execute("""
        SELECT 
            v.name AS volunteer_name,
            v.phone AS volunteer_phone,
            l.name AS leader_name,
            s.name AS section_name,
            q.q1, q.q2, q.q3, q.q4, q.q5, q.q6,
            q.comments, q.created_at
        FROM quality q
        JOIN volunteer v ON q.volunteer_id = v.id
        JOIN leader l ON q.leader_id = l.id
        JOIN section s ON v.section_id = s.id
        ORDER BY q.created_at DESC
    """)
    data1 = cursor.fetchall()
    ws1.append(["اسم المتطوعة", "رقم الجوال", "اسم القائدة", "الركن", "حسن التعامل، وتقديم المساعدة عند الحاجة.", "الالتزام بالحضور والتواجد في الأوقات المحددة بالركن.", "القدرة على العمل الجماعي وحل المشاكل بشكل فعال.", "مستوى المعرفة بالمهام المسندة لهم واستجابة المتطوعين للتوجيهات.", "الالتزام بميثاق الأخلاقي.", "لدى المتطوع صفات قيادية؟", "ملاحظات", "تاريخ"])
    for row in data1:
        ws1.append(list(row.values()))


    # -------------------- 2. قادة الأركان --------------------
    ws2 = wb.create_sheet("قادة الأركان")
    cursor.execute("""
        SELECT 
            dl.name AS evaluator_name,
            l.name AS leader_name,
            l.phone AS leader_phone,
            s.name AS section_name,
            e.q1, e.q2, e.q3, e.q4, e.q5, e.q6, e.q7, e.q8, e.q9,
            e.comments, e.created_at
        FROM leader_evaluations e
        JOIN departmentleader dl ON e.evaluator_id = dl.id
        JOIN leader l ON e.leader_id = l.id
        JOIN section s ON l.section_id = s.id
        ORDER BY e.created_at DESC
    """)
    data2 = cursor.fetchall()
    ws2.append(["قائد القسم", "قائد الركن", "رقم الجوال", "الركن", "مدى التزام القائد بالحضور في الوقت المحدد", "تعامل القائد مع المتطوعات باحترام ولباقة", "قدرة القائد على حل المشكلات الميدانية", "وضوح القائد في شرح المهام للمتطوعات", "سرعة استجابة القائد للتوجيهات والتعليمات", "تنظيم القائد للمهام اليومية", "متابعة القائد لأداء المتطوعات", "تشجيع القائد للمتطوعات وتحفيزهن", "مدى مساهمة القائد في تحسين بيئة العمل", "ملاحظات", "تاريخ"])
    for row in data2:
        ws2.append(list(row.values()))

    # -------------------- 3. متطوعي اللجان --------------------
    ws3 = wb.create_sheet("متطوعي اللجان")
    cursor.execute("""
        SELECT 
            cl.name AS evaluator_name,
            cv.name AS volunteer_name,
            cv.phone AS volunteer_phone,
            cl.committee_section AS section_name,
            e.q1, e.q2, e.q3, e.q4, e.q5, e.q6, e.q7, e.q8, e.q9, e.q10,
            e.comments, e.created_at
        FROM committee_volunteer_evaluations e
        JOIN committeeleader cl ON e.evaluator_id = cl.id
        JOIN committeevolunteer cv ON e.volunteer_id = cv.id
        ORDER BY e.created_at DESC
    """)
    data3 = cursor.fetchall()
    ws3.append(["قائدة اللجنة", "المتطوعة", "رقم الجوال", "اللجنة", "ما مدى التزام المتطوعة بالحضور والانصراف في الوقت المحدد؟", "ما مدى التزام المتطوعة بالمهام الموكلة إليها؟", "ما مدى تعاون المتطوعة مع بقية الفريق داخل اللجنة؟", "ما مدى احترام المتطوعة للتعليمات والتوجيهات الصادرة من القائدة؟", "ما مدى مرونة المتطوعة في التعامل مع التغييرات أو الظروف الطارئة؟", "ما مدى حُسن تواصل المتطوعة مع القائدة وبقية الفريق؟", "ما مدى حماس ومبادرة المتطوعة في أداء مهامها؟", "ما مدى التزام المتطوعة بالزي المناسب والسلوكيات العامة؟", "ما مدى قدرة المتطوعة على أداء المهام بجودة واهتمام؟", "ما مدى التزام المتطوعة بالأخلاقيات واحترام الآخرين أثناء العمل؟", "ملاحظات", "تاريخ"])
    for row in data3:
        ws3.append(list(row.values()))

        
    # حفظ الملف في الذاكرة
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="تقرير_الجودة.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/download_transfer_requests")
def download_transfer_requests():
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "طلبات النقل"

    # جلب البيانات مع ربط أسماء القادة والمتطوعات والأركان
    cursor.execute("""
        SELECT 
            r.id,
            l.name AS leader_name,
            v.name AS volunteer_name,
            r.request_date,
            s1.name AS from_section,
            s2.name AS to_section,
            r.status,
            r.note
        FROM request r
        JOIN leader l ON r.leader_id = l.id
        JOIN volunteer v ON r.volunteer_id = v.id
        JOIN section s1 ON r.from_section_id = s1.id
        JOIN section s2 ON r.to_section_id = s2.id
        ORDER BY r.request_date DESC
    """)
    data = cursor.fetchall()

    # إضافة العناوين
    ws.append(["رقم الطلب", "اسم القائدة", "اسم المتطوعة", "تاريخ الطلب", "من ركن", "إلى ركن", "الحالة", "ملاحظة"])

    # تعبئة البيانات
    for row in data:
        ws.append(list(row.values()))

    # إنشاء الملف وتحميله
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="طلبات_النقل.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/download_attendance_report")
def download_attendance_report():
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    wb = Workbook()

    # -------------------- 1. تحضير الميدان --------------------
    ws1 = wb.active
    ws1.title = "تحضير الميدان"

    # جلب المتطوعات مع أركانهن
    cursor.execute("""
        SELECT v.id, v.name, v.phone, s.name AS section_name
        FROM volunteer v
        JOIN section s ON v.section_id = s.id
    """)
    volunteers = cursor.fetchall()

    # جلب التواريخ
    cursor.execute("SELECT DISTINCT attendance_date FROM volunteerattendance ORDER BY attendance_date")
    dates = [row["attendance_date"] for row in cursor.fetchall()]

    # العناوين
    header = ["الاسم", "الجوال", "الركن"] + [d.strftime("%Y-%m-%d") for d in dates]
    ws1.append(header)

    for v in volunteers:
        row = [v["name"], v["phone"], v["section_name"]]
        for date in dates:
            cursor.execute("""
                SELECT 
                    COALESCE(l.name, dl.name) AS preserver_name
                FROM volunteerattendance va
                LEFT JOIN leader l ON va.leader_id = l.id
                LEFT JOIN departmentleader dl ON va.department_leader_id = dl.id
                WHERE va.volunteer_id = %s AND va.attendance_date = %s
            """, (v["id"], date))
            result = cursor.fetchone()
            row.append(result["preserver_name"] if result else "")
        ws1.append(row)

    # -------------------- 2. تحضير اللجان --------------------
    ws2 = wb.create_sheet("تحضير اللجان")

    cursor.execute("""
        SELECT cv.id, cv.name, cv.phone, cv.section
        FROM committeevolunteer cv
    """)
    committee_vols = cursor.fetchall()

    cursor.execute("SELECT DISTINCT attendance_date FROM committeeattendance ORDER BY attendance_date")
    committee_dates = [row["attendance_date"] for row in cursor.fetchall()]

    header2 = ["الاسم", "الجوال", "اللجنة"] + [d.strftime("%Y-%m-%d") for d in committee_dates]
    ws2.append(header2)

    for v in committee_vols:
        row = [v["name"], v["phone"], v["section"]]
        for date in committee_dates:
            cursor.execute("""
                SELECT l.name
                FROM committeeattendance ca
                JOIN committeeleader l ON ca.leader_id = l.id
                WHERE ca.volunteer_id = %s AND ca.attendance_date = %s
            """, (v["id"], date))
            result = cursor.fetchone()
            row.append(result["name"] if result else "")
        ws2.append(row)

    # -------------------- حفظ الملف --------------------
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="تقرير_تحضير_المتطوعات.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/add_leader", methods=["POST"])
def add_leader():
    leader_id = request.form.get("id")
    name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    section_id = request.form.get("section_id")

    if not (leader_id and name and phone and section_id):
        flash("⚠️ تأكد من تعبئة جميع الحقول لقائد الركن، بما في ذلك المعرف (ID).", "danger")
        return redirect(url_for("hr_dashboard"))

    try:
        cursor.execute("""
            INSERT INTO Leader (id, name, phone, section_id)
            VALUES (%s, %s, %s, %s)
        """, (leader_id, name, phone, section_id))
        db.commit()
        flash("✅ تم إضافة قائد الركن بنجاح.", "success")
    except Exception as e:
        db.rollback()
        flash("❌ حدث خطأ أثناء إضافة قائد الركن: " + str(e), "danger")

    return redirect(url_for("hr_dashboard"))



@app.route("/add_department_leader", methods=["POST"])
def add_department_leader():
    dept_leader_id = request.form.get("id")
    name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    section_ids = request.form.getlist("section_ids")

    if not (dept_leader_id and name and phone and section_ids):
        flash("⚠️ تأكد من تعبئة جميع الحقول لقائد القسم وتحديد الأركان.", "danger")
        return redirect(url_for("hr_dashboard"))

    try:
        cursor.execute("""
            INSERT INTO departmentleader (id, name, phone)
            VALUES (%s, %s, %s)
        """, (dept_leader_id, name, phone))

        for section_id in section_ids:
            cursor.execute("""
                UPDATE section
                SET department_leader_id = %s
                WHERE id = %s
            """, (dept_leader_id, section_id))

        db.commit()
        flash("✅ تم إضافة قائد القسم وربطه بالأركان بنجاح.", "success")
    except Exception as e:
        db.rollback()
        flash("❌ حدث خطأ أثناء إضافة قائد القسم: " + str(e), "danger")

    return redirect(url_for("hr_dashboard"))


@app.route("/add_committee_leader", methods=["POST"])
def add_committee_leader():
    leader_id = request.form.get("id")
    name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    committee_section = request.form.get("committee_section", "").strip()

    if not (leader_id and name and phone and committee_section):
        flash("⚠️ تأكد من تعبئة جميع الحقول لقائد اللجنة، بما في ذلك المعرف (ID).", "danger")
        return redirect(url_for("hr_dashboard"))

    try:
        cursor.execute("""
            INSERT INTO committeeleader (id, name, phone, committee_section, created_at)
            VALUES (%s, %s, %s, %s, NOW())
        """, (leader_id, name, phone, committee_section))
        db.commit()
        flash("✅ تم إضافة قائد اللجنة بنجاح.", "success")
    except Exception as e:
        db.rollback()
        flash("❌ حدث خطأ أثناء إضافة قائد اللجنة: " + str(e), "danger")

    return redirect(url_for("hr_dashboard"))
@app.route("/add_section", methods=["POST"])
def add_section():
    name = request.form.get("name", "").strip()
    min_val = request.form.get("min")
    max_val = request.form.get("max")

    if not (name and min_val and max_val):
        flash("⚠️ تأكد من تعبئة جميع الحقول لإضافة الركن.", "danger")
        return redirect(url_for("hr_dashboard"))

    try:
        cursor.execute("""
            INSERT INTO section (name, min, max)
            VALUES (%s, %s, %s)
        """, (name, min_val, max_val))
        db.commit()
        flash("✅ تم إضافة الركن بنجاح.", "success")
    except Exception as e:
        db.rollback()
        flash("❌ حدث خطأ أثناء الإضافة: " + str(e), "danger")

    return redirect(url_for("hr_dashboard"))

@app.route("/delete_section", methods=["POST"])
def delete_section():
    section_id = request.form.get("section_id")
    try:
        section_id = int(section_id)

        # حذف القادة المرتبطين بهذا الركن
        cursor.execute("DELETE FROM Leader WHERE section_id = %s", (section_id,))
        
        # حذف الركن نفسه
        cursor.execute("DELETE FROM section WHERE id = %s", (section_id,))
        
        db.commit()
        flash("🗑️ تم حذف الركن وجميع القادة المرتبطين به.", "success")
    except Exception as e:
        db.rollback()
        flash("❌ حدث خطأ أثناء الحذف: " + str(e), "danger")

    return redirect(url_for("hr_dashboard"))


@app.route("/download_exhibition_candidates")
def download_exhibition_candidates():
    import pandas as pd
    import io
    from flask import send_file, flash, redirect, url_for

    # جلب تواريخ الحملة الأحدث
    cursor.execute("SELECT campaign_start_date, campaign_end_date FROM campaigndates ORDER BY id DESC LIMIT 1")
    campaign = cursor.fetchone()
    if not campaign:
        flash("⚠️ لم يتم تحديد تواريخ الحملة بعد.", "warning")
        return redirect(url_for("hr_dashboard"))

    start_date = campaign["campaign_start_date"]
    end_date = campaign["campaign_end_date"]
    total_days = (end_date - start_date).days + 1

    # جلب بيانات المتطوعات
    cursor.execute("""
        SELECT v.id, v.name, v.phone, v.section_id, v.notes, s.name AS section_name
        FROM volunteer v
        JOIN section s ON v.section_id = s.id
    """)
    volunteers = cursor.fetchall()

    # بناء بيانات المرشحين
    all_data = []
    for vol in volunteers:
        cursor.execute("""
            SELECT attendance_date, status, leader_id, department_leader_id
            FROM volunteerattendance
            WHERE volunteer_id = %s
            ORDER BY attendance_date
        """, (vol["id"],))
        records = cursor.fetchall()

        present_days = sum(
            1 for r in records 
            if r["status"] == "✔" and (r["leader_id"] or r["department_leader_id"])
        )

        attendance_percentage = round(present_days / total_days * 100, 1) if total_days else 0

        dates = [r["attendance_date"] for r in records if r["status"] != "present"]
        four_consecutive_absence = False
        if len(dates) >= 4:
            for i in range(len(dates) - 3):
                if (dates[i+3] - dates[i]).days == 3:
                    four_consecutive_absence = True
                    break

        passed_attendance = attendance_percentage >= 70
        passed_absence = not four_consecutive_absence
        passed_all = passed_attendance and passed_absence

        all_data.append({
            "الاسم": vol["name"],
            "رقم الجوال": vol["phone"],
            "الركن": vol["section_name"],
            "عدد أيام الحضور": present_days,
            "نسبة الحضور": f"{attendance_percentage}%",
            "اجتاز نسبة الحضور؟": "نعم" if passed_attendance else "لا",
            "تغيب 4 أيام متتالية؟": "نعم" if four_consecutive_absence else "لا",
            "مؤهل للترشيح؟": "نعم" if passed_all else "لا",
            "ملاحظات": vol["notes"] or ""
        })

    # إنشاء ملف Excel
    df = pd.DataFrame(all_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="المرشحات")
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="exhibition_candidates.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/upload_volunteers_excel", methods=["POST"])
def upload_volunteers_excel():
    if "file" not in request.files:
        flash("⚠️ لم يتم تحميل ملف.", "danger")
        return redirect(url_for("hr_dashboard"))

    file = request.files["file"]
    if file.filename == "":
        flash("⚠️ اسم الملف فارغ.", "danger")
        return redirect(url_for("hr_dashboard"))

    try:
        df = pd.read_excel(file)

        # التأكد من وجود الأعمدة المطلوبة
        expected_columns = {"name", "phone", "section"}
        if not expected_columns.issubset(df.columns.str.lower()):
            flash("⚠️ يجب أن يحتوي الملف على أعمدة: name, phone, section", "danger")
            return redirect(url_for("hr_dashboard"))

        # تأكيد الحقول بصيغة مناسبة
        df.columns = [col.strip().lower() for col in df.columns]

        added_count = 0
        for _, row in df.iterrows():
            name = str(row["name"]).strip()
            phone = str(row["phone"]).strip()
            section_name = str(row["section"]).strip()

            if not (name and phone and section_name):
                continue  # تخطي الصفوف الناقصة

            # جلب section_id من جدول section
            cursor.execute("SELECT id FROM section WHERE name = %s", (section_name,))
            section = cursor.fetchone()
            if not section:
                continue  # تخطي إذا ما وجد القسم

            section_id = section["id"]

            # إدخال المتطوعة
            cursor.execute("""
                INSERT INTO volunteer (name, phone, section_id, number_of_requests, notes)
                VALUES (%s, %s, %s, 0, '')
            """, (name, phone, section_id))
            added_count += 1

        db.commit()
        flash(f"✅ تم رفع الملف وإضافة {added_count} متطوعة.", "success")

    except Exception as e:
        db.rollback()
        flash("❌ حدث خطأ أثناء قراءة الملف: " + str(e), "danger")

    return redirect(url_for("hr_dashboard"))



# =============================
# تشغيل التطبيق
# =============================
if __name__ == "__main__":
    app.run(debug=True)
